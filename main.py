import pandas as pd
import numpy as np
from math import floor
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import xlwings as xw
import schedule_function as sch

import streamlit as st

# タイトルの設定
st.title('初回全体スケジュール調整')

st.header("実行前の確認事項")
# 説明文の設定
st.write('/1_スケジュール実行/inputのなかに、「input.xlsx」という名前のファイルが入っているかを確認してください。')
st.write('確認できたら、実行ボタンを押してください。')


# 「処理を実行」ボタンの表示
if st.button('処理を実行'):
    st.write('ボタンが押されました。ここで処理を開始します。')
    # ここに実行したい処理を記述します




    # 入力のファイルパス
    first_input_file_path = "./1_スケジュール実行/input/input.xlsx"

    # 出力のファイルパス
    first_output_file_path = "./1_スケジュール実行/output/output.xlsx"
    individual_students_schedule_file_path = './2_スケジュール表作成/Data/individual_students_schedule.xlsx'
    individual_subjects_file_path = './2_スケジュール表作成/Data/individual_subjects.xlsx'
    seat_schedule_file_path = "./3_座席表作成/Data/Individual_seat_table.xlsx"

    # ファイルの読み込み
    student_subjects_df = pd.read_excel(first_input_file_path, sheet_name="生徒_受講コマ数").fillna(0)
    teacher_schedule_df = pd.read_excel(first_input_file_path, sheet_name="先生スケジュール").fillna(0)
    student_schedule_df = pd.read_excel(first_input_file_path, sheet_name="生徒スケジュール").fillna(0)
    seat_schedule_update = pd.read_excel(first_input_file_path, sheet_name="座席上限").fillna(0)

    # 生徒の優先順位の情報を追加して辞書形式で作成
    student_info_dict_with_priority = {}

    for index, row in student_subjects_df.iterrows():
        student_name = row["生徒の名前"]
        student_info_dict_with_priority[student_name] = {
            "優先順位": row["生徒の優先順位"],
            "学年": row["学年"],
            "科目": {
                "国語": row["国語"],
                "数学": row["数学"],
                "英語": row["英語"],
                "理科": row["理科"],
                "社会": row["社会"]
            },
            "希望先生": {
                "国語": [row["国語の第1希望の先生"], row["国語の第2希望の先生"]],
                "数学": [row["数学の第1希望の先生"], row["数学の第2希望の先生"]],
                "英語": [row["英語の第1希望の先生"], row["英語の第2希望の先生"]],
                "理科": [row["理科の第1希望の先生"], row["理科の第2希望の先生"]],
                "社会": [row["社会の第1希望の先生"], row["社会の第2希望の先生"]]
            }
        }


    # 「生徒スケジュール確定版」データフレームの作成
    student_schedule_final = student_schedule_df.copy().astype(int)
    student_schedule_final.iloc[:, 2:] = 0

    # 「生徒スケジュール更新用」データフレームの作成
    student_schedule_update = student_schedule_df.copy().astype(int)

    # 「先生スケジュール更新用」データフレームの作成
    teacher_schedule_update = teacher_schedule_df.copy().astype(int)
    teacher_schedule_update["未定"] = 99



    st.write("データの読み込みが完了しました。")
    st.write("スケジューリングを実行します。")



    student_name_list = sorted(student_info_dict_with_priority, key=lambda x: student_info_dict_with_priority[x]['優先順位'])
    subject_name_list = ["国語", "数学", "英語", "理科", "社会"]




    for student_name in student_name_list:
        st.write("----------{}のスケジュール----------".format(student_name))

        # schedule_table_by_dateの初期化
        unique_days = teacher_schedule_update['日付'].drop_duplicates().values

        # 各一意の日付に対して、1から5の限数を持つ新しいカラムを作成
        schedule_table_by_date = pd.DataFrame({
            '日付': np.repeat(unique_days, 5),  # 各日付を5回繰り返し
            '可能コマ数': [0] * len(unique_days) * 5,
            '授業可能日': [0] * len(unique_days) * 5,
            '授業を入れた日付': [0] * len(unique_days) * 5,
        })

        for subject_name in subject_name_list:

            first = sch.check_first_choice_teacher(student_name, subject_name, student_subjects_df)
            second = sch.check_second_choice_teacher(student_name, subject_name, student_subjects_df)
            num_classes = sch.get_number_of_classes(student_name, subject_name, student_info_dict_with_priority)
            st.write("first:{},second:{},num_classes:{}".format(first,second,num_classes))

            if num_classes == 0:
                continue
            
            if first == 1:
                choice = "first"
                subject_table, interval_classes = sch.calculate_schedule_possibility(student_name, subject_name, choice, schedule_table_by_date, num_classes, seat_schedule_update, student_subjects_df, teacher_schedule_df, teacher_schedule_update, student_schedule_update)
                student_schedule_final, student_schedule_update, teacher_schedule_update, num_classes, schedule_table_by_date, seat_schedule_update = sch.process_final_schedule(student_name, subject_name, choice, student_subjects_df, student_schedule_final, student_schedule_update, teacher_schedule_update, subject_table, interval_classes, num_classes, schedule_table_by_date, seat_schedule_update)

                if num_classes == 0:
                    continue

            if num_classes > 0 and second == 1:
                choice = "second"
                subject_table, interval_classes = sch.calculate_schedule_possibility(student_name, subject_name, choice, schedule_table_by_date, num_classes, seat_schedule_update, student_subjects_df, teacher_schedule_df, teacher_schedule_update, student_schedule_update)
                student_schedule_final, student_schedule_update, teacher_schedule_update, num_classes, schedule_table_by_date, seat_schedule_update = sch.process_final_schedule(student_name, subject_name, choice, student_subjects_df, student_schedule_final, student_schedule_update, teacher_schedule_update, subject_table, interval_classes, num_classes, schedule_table_by_date, seat_schedule_update)

                if num_classes == 0:
                    continue

            while num_classes > 0:
                choice = "undecided"
                subject_table, interval_classes = sch.calculate_schedule_possibility(student_name, subject_name, choice, schedule_table_by_date, num_classes, seat_schedule_update, student_subjects_df, teacher_schedule_df, teacher_schedule_update, student_schedule_update)
                student_schedule_final, student_schedule_update, teacher_schedule_update, num_classes, schedule_table_by_date, seat_schedule_update = sch.process_final_schedule(student_name, subject_name, choice, student_subjects_df, student_schedule_final, student_schedule_update, teacher_schedule_update, subject_table, interval_classes, num_classes, schedule_table_by_date, seat_schedule_update)


    # outputファイルの作成
    with pd.ExcelWriter(first_output_file_path) as writer:
        student_schedule_final.to_excel(writer, sheet_name="全体スケジュール", index=False)
        teacher_schedule_update.to_excel(writer, sheet_name='先生スケジュール', index=False)
        student_schedule_update.to_excel(writer, sheet_name='生徒スケジュール', index=False)
        seat_schedule_update.to_excel(writer, sheet_name='座席上限', index=False)



    # 座席表の処理
    # 入力データのパス
    input_path = student_schedule_final

    # 新しいワークブックを作成
    workbook = openpyxl.Workbook()

    # デフォルトのシートを削除
    std = workbook['Sheet']
    workbook.remove(std)

    # ユニークな日付を取得
    unique_dates = input_path['日付'].unique()

    # ユニークな日付をイテレートし、変換関数を適用
    for date in unique_dates:
        date_data = input_path[input_path['日付'] == date]
        transformed_data_date = sch.transform_day_data(date_data).replace(0, np.nan)  # 0の値をNaNに置き換える

        # 新しいシートを作成
        sheet = workbook.create_sheet(title=f"日付_{date}")

        # DataFrameをシートに書き込む
        for i in dataframe_to_rows(transformed_data_date, index=False, header=True):
            sheet.append(i)

    # ワークブックを保存
    workbook.save(seat_schedule_file_path)


    # 個別日程表の作成
    # 科目のコマ数だけを取り出す
    subject_columns = ["国語のコマ数", "数学のコマ数", "英語のコマ数", "理科のコマ数", "社会のコマ数"]

    writer_students = pd.ExcelWriter(individual_students_schedule_file_path, engine='openpyxl')
    writer_subjects = pd.ExcelWriter(individual_subjects_file_path, engine='openpyxl')

    # 生徒ごとに処理
    for student_name in student_name_list:
        # 生徒のスケジュールを取り出す
        student_schedule_data = student_schedule_final[['日付', '限数', student_name]].copy()
        student_schedule_data[student_name] = student_schedule_data[student_name].replace(0, '')
        student_schedule_data['科目'] = student_schedule_data[student_name].astype(str).apply(lambda x: x.split('・')[0] if '・' in x else x)
        student_schedule_data['先生'] = student_schedule_data[student_name].astype(str).apply(lambda x: x.split('・')[1] if '・' in x else '')
        student_schedule_data.drop(columns=[student_name], inplace=True)
        student_schedule_data["生徒名"] = student_name
        # 科目ごとにコマ数を取得し、student_schedule_dataに追加
        for subject in subject_name_list:
            student_schedule_data[subject + "のコマ数"] = student_info_dict_with_priority[student_name]["科目"][subject]
        
        subject_df = student_schedule_data[subject_columns]

        # シートに書き込む
        student_schedule_data.to_excel(writer_students, index=False, sheet_name=student_name)
        subject_df.to_excel(writer_subjects, index=False, sheet_name=student_name)

    # ファイルを保存
    writer_students.save()
    writer_subjects.save()



    st.write("スケジューリングが完成しました。")
